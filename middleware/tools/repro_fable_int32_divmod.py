#!/usr/bin/env python3
"""Minimal repro: fable.Int32 breaks openpyxl column letters (arctrl 3.2 Write).

Run:
    uv run --with 'arctrl==3.2.1' python repro_fable_int32_divmod.py
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from arctrl import ARC  # type: ignore[import-untyped]
from fable_library.core import int32  # type: ignore[import-untyped]
from openpyxl.utils.cell import get_column_letter  # type: ignore[import-untyped]


def repro_divmod() -> None:
    """Reproduce builtins.divmod failing on fable.Int32 (openpyxl column-letter path)."""
    # openpyxl.utils.cell.get_column_letter does: divmod(col_idx, 26)
    # Column indices from FsSpreadsheet/arctrl 3.2 are fable.Int32.
    print("divmod(int32(44), 26) ->", end=" ")
    print(divmod(int32(44), 26))


def repro_openpyxl() -> None:
    """Reproduce openpyxl get_column_letter failing when given fable.Int32."""
    print("get_column_letter(int32(44)) ->", end=" ")
    print(get_column_letter(int32(44)))


def repro_arc_write(rocrate_path: str, out_dir: str) -> None:
    """Write an ARC from a RO-Crate that includes study/assay sheets."""
    arc = ARC.from_rocrate_json_string(Path(rocrate_path).read_text(encoding="utf-8"))
    print(f"ARC.Write({out_dir!r}) ...")
    arc.Write(out_dir)
    print("Write OK")


if __name__ == "__main__":
    print("arctrl / fable-library Int32 × openpyxl divmod repro\n")
    try:
        import importlib.metadata as md

        print("arctrl", md.version("arctrl"))
        print("fable-library", md.version("fable-library"))
        print("openpyxl", md.version("openpyxl"))
        print()
    except Exception as exc:  # noqa: BLE001
        print("version lookup failed:", exc)

    try:
        repro_divmod()
    except TypeError as exc:
        print("TypeError:", exc)

    try:
        repro_openpyxl()
    except TypeError as exc:
        print("TypeError:", exc)

    if len(sys.argv) > 1:
        with tempfile.TemporaryDirectory() as tmp:
            try:
                repro_arc_write(sys.argv[1], tmp)
            except Exception as exc:  # noqa: BLE001
                print("ARC.Write failed:", exc)
                raise SystemExit(1) from exc
